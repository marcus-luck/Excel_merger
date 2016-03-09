from openpyxl import load_workbook
import os
import errno


def main():
    folder = 'input/'
    out_folder = 'output/'

    files = os.listdir(folder)

    # Open old file
    out_name = 'Purchase.xlsx'
    output_name = out_folder + out_name
    print(output_name)

    try:
        wb = open(output_name)
    except IOError as e:
        if e.errno == errno.EACCES:
            return "some default data"
            # Not a permission error.
        raise
    else:
        wb = load_workbook(output_name)

    # Read each file as a matrix and append the content to a list
    for file in files:
        if file[0] is '~':
            pass
        else:
            print(file)
            wb_in = load_workbook(folder + file)
            ws = wb.create_sheet()
            ws.title = file[0:30]
            ws_in = wb_in.active

            for row in ws_in.iter_rows():
                sb = []
                for cell in row:
                    if cell.value is not None:
                        sb.append(str(cell.value))
                    else:
                        sb.append('')

                ws.append(sb)

    output_name = out_folder + out_name
    print()

    try:
        wb.save(output_name)
    except:
        print("Please close Excel file")

    # Move read files to input/old
    old_input = folder + 'old/'
    if os.path.isdir(old_input) is False:
        os.makedirs(old_input)
        for file in files:
            os.rename(folder + file, old_input + file)
    else:
        for file in files:
            os.rename(folder + file, old_input + file)

if __name__ == '__main__':
    main()

